#!/usr/bin/env python3
"""
Generate the Nova Rollout Draft Timeline v0 Excel file.
Run: python3 docs/roadmap/nova-rollout-timeline-v0.py
Output: docs/roadmap/Nova_Rollout_Timeline_v0.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Draft Timeline v0"

# ── Colors ──────────────────────────────────────────────────────────────
CHARCOAL = "32373C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
BORDER_GRAY = "D9D9D9"
HEADER_BG = "32373C"

# Track colors (very light tints — professional, not childish)
TRACK_COLORS = {
    1: "E8F0FE",  # blue tint — Nova Pipeline
    2: "E6F4EA",  # green tint — Tracking
    3: "FFF8E1",  # yellow tint — Enablement
    4: "F3E8FD",  # purple tint — Integrations
    5: "FFF3E0",  # orange tint — Analytics
}

# Milestone cell fills (for populated week cells)
MILESTONE_FILL = {
    1: "D2E3FC",
    2: "CEEAD6",
    3: "FFEEBA",
    4: "E1D5F0",
    5: "FFE0B2",
}

# ── Fonts ───────────────────────────────────────────────────────────────
title_font = Font(name="Segoe UI", size=18, bold=True, color=CHARCOAL)
subtitle_font = Font(name="Segoe UI", size=11, italic=True, color="737373")
date_font = Font(name="Segoe UI", size=10, color="737373")
header_font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
track_font = Font(name="Segoe UI", size=10, bold=True, color=CHARCOAL)
cell_font = Font(name="Segoe UI", size=9, color="333333")
note_font = Font(name="Segoe UI", size=9, color="555555")
legend_label = Font(name="Segoe UI", size=9, bold=True, color=CHARCOAL)
legend_font = Font(name="Segoe UI", size=9, color="555555")

# ── Borders ─────────────────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)

# ── Column widths ───────────────────────────────────────────────────────
col_widths = {
    "A": 22,   # Track
    "B": 28,   # Workstream
    "C": 22,   # Week 1
    "D": 22,   # Week 2
    "E": 22,   # Week 3
    "F": 22,   # Week 4
    "G": 22,   # Week 5
    "H": 22,   # Week 6
    "I": 18,   # Owner
    "J": 36,   # Dependencies / Notes
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# ── Helper ──────────────────────────────────────────────────────────────
def style_cell(cell, font=cell_font, fill=None, alignment=None, border=thin_border):
    cell.font = font
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    if border:
        cell.border = border

wrap_top = Alignment(wrap_text=True, vertical="top")
center_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ── Title block ─────────────────────────────────────────────────────────
ws.merge_cells("A1:J1")
c = ws["A1"]
c.value = "Nova Rollout — Draft Timeline (v0)"
c.font = title_font
c.alignment = Alignment(vertical="center")

ws.merge_cells("A2:J2")
c = ws["A2"]
c.value = "DRAFT — For PM Review & Collaboration  •  Not a final plan — designed to be discussed, reshuffled, and owned together"
c.font = subtitle_font

ws.merge_cells("A3:J3")
c = ws["A3"]
c.value = "Prepared: April 15, 2026  •  Author: Steven Junop  •  For: Abi (Senior PM)  •  Start: April 21, 2026"
c.font = date_font

ws.row_dimensions[1].height = 32
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 8  # spacer

# ── Headers (row 5) ────────────────────────────────────────────────────
headers = [
    "Track",
    "Workstream",
    "Week 1\nApr 21–25",
    "Week 2\nApr 28–May 2",
    "Week 3\nMay 5–9",
    "Week 4\nMay 12–16",
    "Week 5\nMay 19–23",
    "Week 6\nMay 26–30",
    "Owner",
    "Dependencies / Notes",
]

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
for i, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = thin_border
ws.row_dimensions[5].height = 36

# ── Data rows ───────────────────────────────────────────────────────────
# Format: (track_num, track_name, workstream, w1, w2, w3, w4, w5, w6, owner, notes)
rows_data = [
    # ─── Track 1: Nova Pipeline ─────────────────────────────────────────
    (1, "Nova Pipeline\n(Engineering)", "Internal Validation",
     "Run 2–3 real campaigns\nend-to-end, fix issues", "", "", "", "", "",
     "Steven", "Campaigns selected with Jenn"),
    (1, "", "WP Auto-Publish",
     "", "Wire WP MCP into Stage 1\nPer-location pages + URL capture", "", "", "", "",
     "Steven", "WP credentials secured"),
    (1, "", "Auto UTM Linking",
     "", "WP publish → auto-create\ntracked links → notify recruiter", "", "", "", "",
     "Steven", "Depends on WP auto-publish"),
    (1, "", "Organic Content",
     "", "", "Social post copy (LI, FB, X)\nFlyer + poster + social card formats", "", "", "",
     "Steven", "Different tone from paid — needs review"),
    (1, "", "Video Pipeline",
     "", "", "Kling Stage 5 on real campaigns", "", "", "",
     "Steven", "Kling credits needed"),
    (1, "", "Polish & Stabilize",
     "", "", "", "Bug fixes from pilot feedback\n3–5 campaigns fully validated", "", "",
     "Steven", ""),

    # ─── Track 2: Tracking & Attribution ────────────────────────────────
    (2, "Tracking &\nAttribution\n(GTM / GA4)", "GTM Bug Fixes",
     "Fix 5 existing bugs\n(utm typo, property ID,\nconversions, survey params)", "", "", "", "", "",
     "Steven + Poola", "GTM publish access required"),
    (2, "", "New Event Tags",
     "7 new tags + UTM persist\n+ visitor ID + cross-domain", "", "", "", "", "",
     "Steven + Poola", "Coordinate with Poola"),
    (2, "", "GA4 Dimensions",
     "", "12 custom dimensions\n+ initial QA", "", "", "", "",
     "Steven", "GA4 admin access required"),
    (2, "", "Reporting Dashboard",
     "", "", "Looker Studio dashboard\nVerify data flow", "", "", "",
     "Steven", "Needs 1–2 weeks of clean data"),

    # ─── Track 3: Team Enablement ───────────────────────────────────────
    (3, "Team\nEnablement\n(Recruiters /\nDesigners)", "Designer Testing",
     "Miguel: daily portal use\non real campaigns", "Ongoing — feedback on\ngallery + edit tools", "", "", "", "",
     "Miguel", ""),
    (3, "", "Internal Pilot",
     "", "Steven + Jenn run 2–3\ncampaigns, document pain points", "", "", "", "",
     "Steven + Jenn", ""),
    (3, "", "Recruiter Pilot",
     "", "", "1–2 recruiter volunteers\nsupervised intake walkthrough", "", "", "",
     "Jenn (recruit)\nSteven (support)", "Jenn to identify volunteers"),
    (3, "", "Training & Docs",
     "", "", "Record Loom walkthroughs\nfor each portal", "Full recruiter team\ntraining session", "", "",
     "Steven", ""),
    (3, "", "Agency Handoff",
     "", "", "", "First real magic link\nhandoff — gather feedback", "", "",
     "Steven + Jenn", "Need agency contact info"),

    # ─── Track 4: Integrations ──────────────────────────────────────────
    (4, "Integrations\n(Microsoft\nEcosystem)", "IT Request Submission",
     "Send integration reqs\nto IT (SSO, Teams,\nOutlook, SharePoint)", "", "", "", "", "",
     "Steven", "P1 — unblocks everything"),
    (4, "", "Teams Webhook (Prod)",
     "", "Get production webhook\nURL from IT", "", "", "", "",
     "IT Admin", "Currently using test URL"),
    (4, "", "Outlook / Graph API",
     "", "App registration for\n@centific.com sending", "", "", "", "",
     "IT Admin", "Currently using test setup"),
    (4, "", "Azure AD SSO",
     "", "", "Clerk SAML ↔ Azure AD\nEnterprise app config", "", "", "",
     "IT Admin + Steven", "Blocks internal rollout"),
    (4, "", "SharePoint (P2)",
     "", "", "", "Auto-create campaign\nfolders, save assets", "", "",
     "IT Admin + Steven", "Nice to have — not blocking"),

    # ─── Track 5: Analytics / Reporting ─────────────────────────────────
    (5, "Analytics &\nReporting\n(VYRA Visualize)", "Data Accumulation",
     "", "", "Live campaigns generating\nUTM click data", "Continue accumulating\nproduction data", "", "",
     "—", "Blocked until Track 2 live"),
    (5, "", "API Port",
     "", "", "", "Port SRC Command Center\nAPI to Nova data model", "KPI Dashboard +\nchannel rollups in UI", "",
     "Steven", "Depends on GA4 API access"),
    (5, "", "RevBrain + Funnel",
     "", "", "", "", "RevBrain AI recs\n+ funnel drilldown", "",
     "Steven", "Needs ad platform API creds"),
    (5, "", "Exports & Sharing",
     "", "", "", "", "", "PDF/CSV reports\nPublic share tokens",
     "Steven", "For Adam pitch deck"),
]

# Write data rows
row_num = 6
current_track = None
for (track_n, track_name, workstream, w1, w2, w3, w4, w5, w6, owner, notes) in rows_data:
    track_fill = TRACK_COLORS[track_n]
    ms_fill = MILESTONE_FILL[track_n]

    # Track name (col A)
    c = ws.cell(row=row_num, column=1, value=track_name if track_name else "")
    style_cell(c, font=track_font, fill=track_fill, alignment=wrap_top)

    # Workstream (col B)
    c = ws.cell(row=row_num, column=2, value=workstream)
    style_cell(c, font=cell_font, fill=track_fill)

    # Weeks (cols C-H)
    for col_idx, val in enumerate([w1, w2, w3, w4, w5, w6], 3):
        c = ws.cell(row=row_num, column=col_idx, value=val)
        fill = ms_fill if val.strip() else LIGHT_GRAY
        style_cell(c, font=cell_font, fill=fill)

    # Owner (col I)
    c = ws.cell(row=row_num, column=9, value=owner)
    style_cell(c, font=cell_font, fill=track_fill, alignment=center_wrap)

    # Notes (col J)
    c = ws.cell(row=row_num, column=10, value=notes)
    style_cell(c, font=note_font, fill=WHITE)

    ws.row_dimensions[row_num].height = 48
    row_num += 1

# ── Merge track name cells ──────────────────────────────────────────────
# Find consecutive rows for each track and merge col A
track_ranges = {}
for i, (track_n, track_name, *_) in enumerate(rows_data):
    if track_n not in track_ranges:
        track_ranges[track_n] = {"start": 6 + i, "end": 6 + i, "name": track_name}
    else:
        track_ranges[track_n]["end"] = 6 + i

for track_n, info in track_ranges.items():
    if info["start"] != info["end"]:
        ws.merge_cells(start_row=info["start"], start_column=1,
                       end_row=info["end"], end_column=1)
        c = ws.cell(row=info["start"], column=1)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ── Key Milestones row ──────────────────────────────────────────────────
row_num += 1  # blank spacer
ms_row = row_num
ws.merge_cells(f"A{ms_row}:B{ms_row}")
c = ws.cell(row=ms_row, column=1, value="KEY MILESTONES")
c.font = Font(name="Segoe UI", size=10, bold=True, color=CHARCOAL)
c.alignment = wrap_top

milestones = [
    "GTM bugs fixed\n(Day 1)",
    "WP auto-publish live\nDesigner testing",
    "Recruiter pilot starts\nSSO configured",
    "Full team trained\nAgency test handoff",
    "KPI Dashboard live\nRevBrain active",
    "Adam pitch ready\nExports + share links",
]
for i, m in enumerate(milestones, 3):
    c = ws.cell(row=ms_row, column=i, value=m)
    c.font = Font(name="Segoe UI", size=9, bold=True, color=CHARCOAL)
    c.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    c.alignment = center_wrap
    c.border = thin_border
ws.row_dimensions[ms_row].height = 40

# ── Dependencies / Blockers section ─────────────────────────────────────
row_num = ms_row + 2
ws.merge_cells(f"A{row_num}:J{row_num}")
c = ws.cell(row=row_num, column=1, value="OPEN DEPENDENCIES — Items that need PM/Org action")
c.font = Font(name="Segoe UI", size=11, bold=True, color=CHARCOAL)
row_num += 1

deps = [
    ("P1", "Azure AD SSO configuration", "IT Admin", "Blocks internal team access — using Clerk dev keys"),
    ("P1", "Teams production webhook URL", "IT Admin", "Currently using test webhook"),
    ("P1", "Outlook Graph API app registration", "IT Admin", "Needed for @centific.com notifications"),
    ("P1", "GTM publish access (GTM-NR965959)", "Poola / IT", "Blocks all tracking fixes"),
    ("P1", "GA4 admin access (property 330157295)", "Poola / IT", "Blocks custom dimensions + conversion marking"),
    ("P1", "1–2 recruiter pilot volunteers", "Jenn", "Needed by Week 3"),
    ("P1", "Agency contact for magic link test", "Jenn", "Needed by Week 4"),
    ("P2", "SharePoint site URL + Graph API permissions", "IT Admin", "Nice to have — not blocking launch"),
    ("P2", "Ad platform API creds (Meta, LinkedIn, Google)", "Steven + IT", "Blocks Track 5 analytics"),
]

dep_headers = ["Priority", "Dependency", "Owner", "Impact"]
for i, h in enumerate(dep_headers):
    col = [1, 2, 5, 7][i]
    c = ws.cell(row=row_num, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_wrap
    c.border = thin_border

# Merge header cells for wider columns
ws.merge_cells(f"B{row_num}:D{row_num}")
ws.merge_cells(f"E{row_num}:F{row_num}")
ws.merge_cells(f"G{row_num}:J{row_num}")
row_num += 1

for pri, dep, owner, impact in deps:
    c = ws.cell(row=row_num, column=1, value=pri)
    pri_fill = "FFCDD2" if pri == "P1" else "FFF9C4"
    style_cell(c, fill=pri_fill, alignment=center_wrap)

    ws.merge_cells(f"B{row_num}:D{row_num}")
    c = ws.cell(row=row_num, column=2, value=dep)
    style_cell(c)

    ws.merge_cells(f"E{row_num}:F{row_num}")
    c = ws.cell(row=row_num, column=5, value=owner)
    style_cell(c, alignment=center_wrap)

    ws.merge_cells(f"G{row_num}:J{row_num}")
    c = ws.cell(row=row_num, column=7, value=impact)
    style_cell(c, font=note_font)

    ws.row_dimensions[row_num].height = 28
    row_num += 1

# ── Success Metrics section ─────────────────────────────────────────────
row_num += 1
ws.merge_cells(f"A{row_num}:J{row_num}")
c = ws.cell(row=row_num, column=1, value="SUCCESS METRICS — How we'll know this is working")
c.font = Font(name="Segoe UI", size=11, bold=True, color=CHARCOAL)
row_num += 1

metric_headers = ["Metric", "", "Before Nova", "", "With Nova", "", "How to Measure", "", "", ""]
for i, h in enumerate(metric_headers):
    if h:
        col = [1, 0, 3, 0, 5, 0, 7, 0, 0, 0][i]
        if col:
            c = ws.cell(row=row_num, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_wrap
            c.border = thin_border

ws.merge_cells(f"A{row_num}:B{row_num}")
ws.merge_cells(f"C{row_num}:D{row_num}")
ws.merge_cells(f"E{row_num}:F{row_num}")
ws.merge_cells(f"G{row_num}:J{row_num}")
row_num += 1

metrics = [
    ("Time: JD → campaign package", "3–5 business days", "30 min + review", "Timestamp: intake submit → final approval"),
    ("Creatives per campaign", "2–4 generic designs", "15–30+ per-persona", "Asset count in system"),
    ("Cost per campaign", "Agency hours + designer time", "~$0 (free AI + Vercel)", "Infrastructure cost tracking"),
    ("Recruiter time spent", "4–8 hours coordinating", "15 min intake + 10 min links", "Session tracking"),
    ("Tracking accuracy", "Broken (7,600 paid visitors → 0 conversions)", "Full cross-device attribution", "GA4 conversion reports"),
]

for metric, before, after, measure in metrics:
    ws.merge_cells(f"A{row_num}:B{row_num}")
    c = ws.cell(row=row_num, column=1, value=metric)
    style_cell(c, font=Font(name="Segoe UI", size=9, bold=True, color="333333"))

    ws.merge_cells(f"C{row_num}:D{row_num}")
    c = ws.cell(row=row_num, column=3, value=before)
    style_cell(c, fill="FFEBEE")  # light red

    ws.merge_cells(f"E{row_num}:F{row_num}")
    c = ws.cell(row=row_num, column=5, value=after)
    style_cell(c, fill="E8F5E9")  # light green

    ws.merge_cells(f"G{row_num}:J{row_num}")
    c = ws.cell(row=row_num, column=7, value=measure)
    style_cell(c, font=note_font)

    ws.row_dimensions[row_num].height = 28
    row_num += 1

# ── Footer ──────────────────────────────────────────────────────────────
row_num += 1
ws.merge_cells(f"A{row_num}:J{row_num}")
c = ws.cell(row=row_num, column=1,
            value="This is a working draft. Timelines are estimates — designed to be adjusted together. All tracks can shift based on IT response times and pilot feedback.")
c.font = Font(name="Segoe UI", size=9, italic=True, color="999999")
c.alignment = Alignment(horizontal="center")

# ── Print setup ─────────────────────────────────────────────────────────
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# ── Freeze panes ────────────────────────────────────────────────────────
ws.freeze_panes = "C6"

# ── Save ────────────────────────────────────────────────────────────────
output = "/Users/stevenjunop/centric-intake/docs/roadmap/Nova_Rollout_Timeline_v0.xlsx"
wb.save(output)
print(f"Saved: {output}")
